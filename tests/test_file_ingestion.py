import io
from datetime import datetime

import openpyxl
import pytest

from tools.excel_parser import parse


class TestColumnMapping:
    """Test that non-canonical column headers are mapped correctly."""

    def test_maps_reference_id_variants(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        assert all("referenceId" in r for r in records)

    def test_maps_txn_amount_to_amount(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        assert records[0]["amount"] == 100.00

    def test_maps_txn_status_to_status(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        assert records[0]["status"] == "success"

    def test_maps_created_at_to_timestamp(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        assert isinstance(records[0]["timestamp"], datetime)

    def test_canonical_headers_pass_through(self):
        """Columns already named canonically should work without mapping."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["referenceId", "amount", "status", "timestamp"])
        ws.append(["REF001", 100.0, "success", datetime(2025, 1, 1)])
        buf = io.BytesIO()
        wb.save(buf)

        records, _, _ = parse(buf.getvalue())
        assert len(records) == 1
        assert records[0]["referenceId"] == "REF001"


class TestPeriodInference:
    """Test that billing period is inferred from min/max timestamps."""

    def test_period_inferred_from_timestamps(self, sample_partner_bytes):
        _, period, _ = parse(sample_partner_bytes)
        assert period["start"] == datetime(2025, 1, 5, 10, 0)
        # The empty-referenceId row (Jan 24) is dropped, so max is REF019 (Jan 23)
        assert period["end"] == datetime(2025, 1, 23, 12, 0)

    def test_period_none_when_no_timestamp_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["referenceId", "amount"])
        ws.append(["REF001", 100.0])
        buf = io.BytesIO()
        wb.save(buf)

        _, period, _ = parse(buf.getvalue())
        assert period["start"] is None
        assert period["end"] is None


class TestNullReferenceIdHandling:
    """Test that rows with null/empty referenceId are dropped."""

    def test_drops_empty_reference_id(self, sample_partner_bytes):
        records, _, metadata = parse(sample_partner_bytes)
        ref_ids = [r["referenceId"] for r in records]
        assert "" not in ref_ids
        assert metadata["dropped_rows"] == 1

    def test_drops_null_reference_id(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["referenceId", "amount"])
        ws.append([None, 100.0])
        ws.append(["REF001", 200.0])
        buf = io.BytesIO()
        wb.save(buf)

        records, _, metadata = parse(buf.getvalue())
        assert len(records) == 1
        assert metadata["dropped_rows"] == 1

    def test_metadata_total_rows_excludes_dropped(self, sample_partner_bytes):
        _, _, metadata = parse(sample_partner_bytes)
        assert metadata["total_rows"] == 19  # 20 rows minus 1 empty referenceId


class TestMissingReferenceIdColumn:
    """Test that a descriptive error is raised when referenceId can't be identified."""

    def test_raises_on_unrecognized_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["foo", "bar", "baz"])
        ws.append(["a", "b", "c"])
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="Could not identify referenceId column"):
            parse(buf.getvalue())


class TestHappyPath:
    """Test full parsing with correct output shape."""

    def test_record_count(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        assert len(records) == 19

    def test_record_has_canonical_fields(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        expected_keys = {"referenceId", "amount", "status", "timestamp", "raw"}
        for r in records:
            assert set(r.keys()) == expected_keys

    def test_raw_field_preserved(self, sample_partner_bytes):
        records, _, _ = parse(sample_partner_bytes)
        assert isinstance(records[0]["raw"], dict)
        assert len(records[0]["raw"]) > 0
