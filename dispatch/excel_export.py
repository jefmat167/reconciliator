import os
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from config import settings

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


async def export(state: dict) -> str:
    """Generate the four-sheet reconciliation Excel workbook. Returns 'excel' on success."""
    report = state["report"]
    period = state["period"]

    start_str = period["start"].strftime("%Y-%m-%d") if period.get("start") else "unknown"
    end_str = period["end"].strftime("%Y-%m-%d") if period.get("end") else "unknown"
    filename = f"reconciliation_{start_str}_{end_str}.xlsx"

    os.makedirs(settings.export_output_dir, exist_ok=True)
    filepath = os.path.join(settings.export_output_dir, filename)

    wb = Workbook()

    _write_summary_sheet(wb, report)
    _write_matched_sheet(wb, state["matched_records"])
    _write_missing_sheet(wb, "Missing in Ours", state["missing_in_ours"])
    _write_missing_sheet(wb, "Missing in Partner", state["missing_in_partner"])

    wb.save(filepath)
    logger.info(f"Excel export saved to {filepath}")
    return "excel"


def _style_header_row(ws):
    """Apply bold white text on blue fill to the first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    """Auto-size column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _write_summary_sheet(wb: Workbook, report: dict):
    """Write the Summary sheet with key-value pairs."""
    ws = wb.active
    ws.title = "Summary"

    rows = [
        ("Metric", "Value"),
        ("Period Start", report["period_start"]),
        ("Period End", report["period_end"]),
        ("Partner Total", report["partner_total"]),
        ("Internal Total", report["internal_total"]),
        ("Matched Total", report["matched_total"]),
        ("Matched Clean", report["matched_clean"]),
        ("Matched Flagged", report["matched_flagged"]),
        ("Missing in Ours", report["missing_in_ours_count"]),
        ("Missing in Partner", report["missing_in_partner_count"]),
        ("Match Rate (%)", report["match_rate"]),
        ("Flags", "; ".join(report["flags"]) if report["flags"] else "None"),
        ("Summary", report["summary_text"]),
    ]
    for row in rows:
        ws.append(row)

    _style_header_row(ws)
    _auto_width(ws)


def _write_matched_sheet(wb: Workbook, matched_records: list[dict]):
    """Write the Matched sheet with per-field comparison columns."""
    ws = wb.create_sheet("Matched")

    headers = [
        "referenceId",
        "amount (ours)", "amount (partner)",
        "status (ours)", "status (partner)",
        "match status",
    ]
    ws.append(headers)

    for rec in matched_records:
        partner = rec["partner_data"]
        internal = rec["internal_data"]
        discrepancies = rec["discrepancies"]

        # Build match status string
        if not rec["has_discrepancy"]:
            match_status = "✓ Clean"
        else:
            mismatched_fields = " + ".join(discrepancies.keys())
            match_status = f"⚠ {mismatched_fields} mismatch"

        ws.append([
            rec["referenceId"],
            internal.get("amount"),
            partner.get("amount"),
            internal.get("status"),
            partner.get("status"),
            match_status,
        ])

    _style_header_row(ws)
    _auto_width(ws)

    # Format amount columns as 2 decimal places
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            if cell.value is not None:
                cell.number_format = numbers.FORMAT_NUMBER_00


def _write_missing_sheet(wb: Workbook, title: str, records: list[dict]):
    """Write a missing records sheet."""
    ws = wb.create_sheet(title)

    headers = ["referenceId", "amount", "status", "timestamp"]
    ws.append(headers)

    for rec in records:
        ts = rec.get("timestamp")
        if isinstance(ts, datetime):
            ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
        else:
            ts_str = str(ts) if ts else ""

        ws.append([
            rec.get("referenceId"),
            rec.get("amount"),
            rec.get("status"),
            ts_str,
        ])

    _style_header_row(ws)
    _auto_width(ws)

    # Format amount column as 2 decimal places
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:
                cell.number_format = numbers.FORMAT_NUMBER_00
